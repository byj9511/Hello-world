# from openpyxl import Workbook
#
# wb = Workbook()
# # 创建了第一页
# ws = wb.active
# # 最后一页名为hello
# ws1 = wb.create_sheet('hello')
# # 第一页名为world
# ws2 = wb.create_sheet('world', 0)
# # 第二页名为hate
# ws.title = 'hate'
# # ws与ws3都指向了第二页？
# ws3 = wb['hate']
# print(ws.title)
# print(wb.sheetnames)
# for sheet in wb:
#     print(sheet.title)
# ws4 = wb.copy_worksheet(ws)
# print(ws4.title)
# ws4.title = 'hate_copy'
# print(ws4.title)
#
# ws['A4'] = 5
# ws3['B3'] = 6
# #遍历所有数据
# print(tuple(ws.rows))
# wb.template=True
# wb.save('test.xlsx')
from openpyxl import load_workbook

wb= load_workbook(r'C:\Users\DELL\Desktop\exercise\ZDTX-Load calculation.xlsx',data_only=True)
#help(load_workbook)
sheet_list= wb.get_sheet_names()
sheet2 = wb[sheet_list[1]]
load_combination,load_case=[],[]
#因为sheet.rows是生成器类型，不能使用索引，转换成list之后再使用索引
for cell in list(sheet2.columns)[3]:
    if cell.value:
        load_combination.append(cell.value)
for cell in list(sheet2.rows)[13]:
    # isinstance用元组形式包含多个判断对象
    if isinstance(cell.value,(float,int)):
        load_case.append(cell.value)

#方法2 sheet的切片返回了一个二阶的元祖,所以要用两次for循环
# for row_cell in list(sheet2['D15':'D70']):
#     for cell in row_cell:
#         print(cell.value)
line = ''

for row_index,row_cell in enumerate(list(sheet2['E15':'AW70'])):
    line += 'LOAD COMB %s LOAD COMBINATION %s\n'%(load_combination[row_index],load_combination[row_index])
    for column_index,cell in enumerate(row_cell):
        if cell.value:
            line += str(load_case[column_index])+' '+ str(cell.value) + ' '
    line += '\n'
print(line)
